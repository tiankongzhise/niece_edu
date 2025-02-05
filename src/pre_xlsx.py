import os
import pandas as pd

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import re

def process_merged_cells(file_path):
    """
    处理Excel合并单元格和内容清洗的完整函数
    参数：
        file_path: Excel文件路径
    返回：
        处理后的DataFrame
    """
    
    # ==================== 第一步：解析合并单元格信息 ====================
    wb = load_workbook(filename=file_path, read_only=False) 
    ws = wb.active
    # 存储合并区域信息：{(起始行,起始列): (结束行,结束列)}
    merged_ranges = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        merged_ranges[(min_row-1, min_col-1)] = (max_row-1, max_col-1)  # openpyxl从1开始，pandas从0开始

    # ==================== 第二步：处理合并单元格 ====================
    # 使用pandas读取原始数据
    df = pd.read_excel(file_path, header=None, engine='openpyxl')
    
    # 创建填充矩阵
    fill_matrix = df.copy()
    row_merged_cols = set()  # 记录行合并涉及的列
    col_merged_rows = set()  # 记录列合并涉及的行

    # 遍历所有合并区域
    for (min_row, min_col), (max_row, max_col) in merged_ranges.items():
        # 获取合并单元格的值
        merged_value = df.iloc[min_row, min_col]
        
        # 处理行合并（垂直合并）
        if min_row != max_row and min_col == max_col:
            fill_matrix.iloc[min_row:max_row+1, min_col] = merged_value
            row_merged_cols.add(min_col)
        
        # 处理列合并（水平合并）
        if min_col != max_col and min_row == max_row:
            fill_matrix.iloc[min_row, min_col:max_col+1] = np.nan
            fill_matrix.iloc[min_row, min_col] = merged_value
            col_merged_rows.add(min_row)

    # ==================== 第三步：清理列合并后的空列 ====================
    # 识别需要删除的列（列合并区域右侧的列）
    cols_to_drop = []
    for (min_row, min_col), (max_row, max_col) in merged_ranges.items():
        if min_col != max_col:  # 列合并
            cols_to_drop.extend(range(min_col+1, max_col+1))
    
    # 去重并排序要删除的列
    cols_to_drop = sorted(list(set(cols_to_drop)))
    # 过滤掉超出DataFrame范围的列
    cols_to_drop = [col for col in cols_to_drop if col < df.shape[1]]
    # 删除列
    cleaned_df = fill_matrix.drop(columns=cols_to_drop)

    # ==================== 第四步：清理括号内的换行符 ====================
    def clean_parentheses_content(text):
        if pd.isna(text):
            return text
        # 使用正则表达式替换括号内的换行符和回车符
        return re.sub(
            r'\([^)]*\)', 
            lambda m: m.group(0).replace('\n', ' ').replace('\r', ' '),
            str(text)
        )

    cleaned_df = cleaned_df.applymap(clean_parentheses_content)

    # ==================== 第五步：重置索引和列名 ====================
    # 保留原始索引但重置列名（根据需求调整）
    cleaned_df.columns = range(cleaned_df.shape[1])
    return cleaned_df

def preprocess_excel(input_dir, output_dir):
    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 遍历输入目录中的所有Excel文件
    for filename in os.listdir(input_dir):
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            file_path = os.path.join(input_dir, filename)
            # 读取Excel文件
            processed_df = process_merged_cells(file_path)
            print(processed_df)
            # 在这里进行预处理操作
            # 例如：df = df.dropna()  # 删除空值行
            break
            # 保存预处理后的文件到输出目录
            output_file_path = os.path.join(output_dir, filename)
            # 保存到新文件
            processed_df.to_excel(output_file_path, index=False, header=False)
            print(f"Processed and saved {filename} to {output_file_path}")

if __name__ == "__main__":
    input_directory = './data/excel/'
    output_directory = './data/pre/'
    preprocess_excel(input_directory, output_directory)